# -*- coding: utf-8 -*-
"""
batch_import.py - 批量导入模块
支持 Excel (.xlsx) / CSV 导入，自动识别单片/双胶合/三胶合
"""

import os
import csv
import math
from dataclasses import dataclass, field
from typing import List, Optional, Dict, Any, Tuple

# ── 标准列名（按顺序） ──────────────────────────────────────────────
COLUMN_NAMES = [
    "PartName", "PartNo", "Glass1", "Glass2", "Glass3",
    "T1", "T2", "T3", "R1", "R2", "R3", "R4",
    "MD1", "MD2", "MD3", "AD1", "AD2", "AD3", "AD4",
    "SavePdfFolder", "MfrPdfFolder",
]

REQUIRED_SINGLE  = ["PartName", "PartNo", "Glass1", "T1", "R1", "R2", "MD1", "AD1", "AD2"]
REQUIRED_DOUBLET  = REQUIRED_SINGLE + ["Glass2", "T2", "R3", "MD2", "AD3"]
REQUIRED_TRIPLET  = REQUIRED_DOUBLET + ["Glass3", "T3", "R4", "MD3", "AD4"]


# ── 数据结构 ────────────────────────────────────────────────────────
@dataclass
class SingleLensData:
    """单片镜片数据"""
    glass: str          # 玻璃牌号
    T: float            # 中心厚度
    R_left: float       # 左曲率半径
    R_right: float      # 右曲率半径
    MD: float           # 机械口径
    AD_left: float      # 左口径
    AD_right: float     # 右口径


@dataclass
class CementedLensData:
    """胶合镜片组数据（1~3 片）"""
    part_name: str
    part_no: str
    lenses: List[SingleLensData] = field(default_factory=list)
    save_pdf_folder: str = "Save PDF"
    mfr_pdf_folder: str = "Mfr PDF"
    proc_overrides: Optional[Dict[str, Any]] = None  # 逐行加工参数覆盖 {key: value}
    page_overrides: Optional[Dict[str, Any]] = None  # 逐页加工参数覆盖 {pageIndex: {key: value}}

    @property
    def lens_type(self) -> str:
        n = len(self.lenses)
        if n == 1:   return "单片"
        if n == 2:   return "双胶合"
        if n == 3:   return "三胶合"
        return "未知"

    @property
    def num_pages(self) -> int:
        """PDF 总页数：单片=1，双胶合=3，三胶合=4"""
        n = len(self.lenses)
        return 1 if n == 1 else n + 1

    @property
    def total_thickness(self) -> float:
        return sum(l.T for l in self.lenses)

    @property
    def glass_names_str(self) -> str:
        return "\n".join(l.glass for l in self.lenses if l.glass)


# ── 工具函数 ────────────────────────────────────────────────────────
def _is_empty(val) -> bool:
    if val is None:
        return True
    if isinstance(val, str) and val.strip() == "":
        return True
    if isinstance(val, float) and math.isnan(val):
        return True
    return False


def _safe_float(val) -> Optional[float]:
    if _is_empty(val):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_str(val) -> str:
    if _is_empty(val):
        return ""
    return str(val).strip()


# ── 行解析 ──────────────────────────────────────────────────────────
def parse_row(row_dict: Dict[str, Any]) -> CementedLensData:
    """将一行字典数据解析为 CementedLensData，自动识别胶合类型"""
    part_name = _safe_str(row_dict.get("PartName"))
    part_no   = _safe_str(row_dict.get("PartNo"))
    glass1    = _safe_str(row_dict.get("Glass1"))
    glass2    = _safe_str(row_dict.get("Glass2"))
    glass3    = _safe_str(row_dict.get("Glass3"))

    # 数值参数
    T1  = _safe_float(row_dict.get("T1"))
    T2  = _safe_float(row_dict.get("T2"))
    T3  = _safe_float(row_dict.get("T3"))
    R1  = _safe_float(row_dict.get("R1"))
    R2  = _safe_float(row_dict.get("R2"))
    R3  = _safe_float(row_dict.get("R3"))
    R4  = _safe_float(row_dict.get("R4"))
    MD1 = _safe_float(row_dict.get("MD1"))
    MD2 = _safe_float(row_dict.get("MD2"))
    MD3 = _safe_float(row_dict.get("MD3"))
    AD1 = _safe_float(row_dict.get("AD1"))
    AD2 = _safe_float(row_dict.get("AD2"))
    AD3 = _safe_float(row_dict.get("AD3"))
    AD4 = _safe_float(row_dict.get("AD4"))

    # ── 镜片 1（必填） ──
    miss1 = [k for k in ["T1","R1","R2","MD1","AD1","AD2"]
             if _safe_float(row_dict.get(k)) is None]
    if miss1:
        raise ValueError(f"镜片1缺参数: {', '.join(miss1)}")

    lenses = [SingleLensData(glass=glass1, T=T1, R_left=R1, R_right=R2,
                             MD=MD1, AD_left=AD1, AD_right=AD2)]

    has_g2 = glass2 != "" and T2 is not None
    has_g3 = glass3 != "" and T3 is not None

    # ── 镜片 2（双胶合） ──
    if has_g2:
        miss2 = [k for k in ["T2","R3","MD2","AD3"]
                 if _safe_float(row_dict.get(k)) is None]
        if miss2:
            raise ValueError(f"镜片2缺参数: {', '.join(miss2)}")
        lenses.append(SingleLensData(glass=glass2, T=T2, R_left=R2, R_right=R3,
                                     MD=MD2, AD_left=AD2, AD_right=AD3))

    # ── 镜片 3（三胶合） ──
    if has_g3:
        miss3 = [k for k in ["T3","R4","MD3","AD4"]
                 if _safe_float(row_dict.get(k)) is None]
        if miss3:
            raise ValueError(f"镜片3缺参数: {', '.join(miss3)}")
        lenses.append(SingleLensData(glass=glass3, T=T3, R_left=R3, R_right=R4,
                                     MD=MD3, AD_left=AD3, AD_right=AD4))

    save_folder = _safe_str(row_dict.get("SavePdfFolder", "Save PDF"))
    mfr_folder  = _safe_str(row_dict.get("MfrPdfFolder", "Mfr PDF"))
    if not save_folder:
        save_folder = "Save PDF"
    if not mfr_folder:
        mfr_folder = "Mfr PDF"

    return CementedLensData(
        part_name=part_name, part_no=part_no, lenses=lenses,
        save_pdf_folder=save_folder, mfr_pdf_folder=mfr_folder
    )


# ── 通用行解析 ─────────────────────────────────────────────────────
def _parse_rows(rows: List[tuple]) -> Tuple[List[CementedLensData], List[str]]:
    """将原始行数据解析为 CementedLensData 列表和警告列表。"""
    if len(rows) < 2:
        raise ValueError("至少需要表头行 + 一行数据")

    header = rows[0]
    col_map: Dict[str, int] = {}
    for i, name in enumerate(header):
        if name is None:
            continue
        cn = str(name).strip()
        if cn in COLUMN_NAMES:
            col_map[cn] = i

    missing = [c for c in REQUIRED_SINGLE if c not in col_map]
    if missing:
        raise ValueError(f"缺少必需列: {', '.join(missing)}")

    results: List[CementedLensData] = []
    warnings: List[str] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
            continue
        rd = {}
        for col_name, col_idx in col_map.items():
            if col_idx < len(row):
                rd[col_name] = row[col_idx]
        try:
            results.append(parse_row(rd))
        except ValueError as e:
            warnings.append(f"第 {row_idx} 行: {e}")

    return results, warnings


# ── COM 读取（用于亿赛通等加密环境）────────────────────────────────
def read_excel_via_com(filepath: str) -> Tuple[List[CementedLensData], List[str]]:
    """
    通过 Windows COM 调用 Excel/WPS 应用程序读取文件。
    适用于企业透明加密软件（如亿赛通）管控场景。
    """
    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        # ReadOnly=True 防止与已打开文件的进程冲突
        wb = excel.Workbooks.Open(filepath, ReadOnly=True)
        ws = wb.Worksheets(1)
        used = ws.UsedRange
        rows = []
        for r in range(1, used.Rows.Count + 1):
            row = []
            for c in range(1, used.Columns.Count + 1):
                row.append(ws.Cells(r, c).Value)
            rows.append(tuple(row))
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()
        pythoncom.CoUninitialize()

    return _parse_rows(rows)


# ── Excel 读取 ─────────────────────────────────────────────────────
def read_excel(filepath: str) -> Tuple[List[CementedLensData], List[str]]:
    """
    读取 Excel 文件，返回 (数据列表, 警告列表)。
    若文件被系统加密软件管控（如亿赛通），openpyxl 直接读取会失败，
    此时自动 fallback 到 COM 调用 Excel/WPS 读取。
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as direct_err:
        # openpyxl 失败，尝试 COM 方式（加密环境）
        try:
            return read_excel_via_com(filepath)
        except Exception as com_err:
            raise ValueError(
                f"无法读取 Excel。openpyxl 错误: {direct_err}; COM 错误: {com_err}"
            )

    return _parse_rows(rows)


# ── CSV 读取 ────────────────────────────────────────────────────────
def read_csv_file(filepath: str) -> Tuple[List[CementedLensData], List[str]]:
    """
    读取 CSV 文件，返回 (数据列表, 警告列表)。
    若文件被加密软件管控导致直接读取失败，自动 fallback 到 COM 读取。
    """
    try:
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
    except Exception as direct_err:
        # 直接读取失败，尝试 COM 方式（加密环境）
        try:
            return read_excel_via_com(filepath)
        except Exception as com_err:
            raise ValueError(
                f"无法读取 CSV。直接读取错误: {direct_err}; COM 错误: {com_err}"
            )

    return _parse_rows(rows)


# ── 生成模板 Excel ─────────────────────────────────────────────────
def create_template_excel(filepath: str):
    """生成带示例数据的模板 Excel 文件"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lens Data"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side("thin"), right=Side("thin"),
                  top=Side("thin"), bottom=Side("thin"))

    # 表头
    for i, name in enumerate(COLUMN_NAMES):
        c = ws.cell(row=1, column=i + 1, value=name)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = thin

    # 示例 - 单片
    ex1 = ["singlet_01","100.2.00001","H-K9L","","",
           4.70,"","",-35.0,30.0,"","",
           19.0,"","",15.0,13.0,"","",
           "Save PDF","Mfr PDF"]
    # 示例 - 双胶合
    ex2 = ["doublet_01","100.2.00002","H-K9L","ZF2","",
           3.50,2.00,"",-50.0,-35.0,60.0,"",
           20.0,18.0,"",16.0,15.0,14.0,"",
           "Save PDF","Mfr PDF"]
    # 示例 - 三胶合
    ex3 = ["triplet_01","100.2.00003","H-K9L","ZF2","H-K9L",
           4.00,2.50,3.00,-60.0,-40.0,45.0,80.0,
           22.0,20.0,21.0,18.0,17.0,16.0,15.0,
           "Save PDF","Mfr PDF"]

    for r, ex in enumerate([ex1, ex2, ex3], start=2):
        for i, val in enumerate(ex):
            c = ws.cell(row=r, column=i + 1, value=val if val != "" else None)
            c.border = thin; c.alignment = Alignment(horizontal="center")

    widths = [14,14,10,10,10,8,8,8,8,8,8,8,8,8,8,8,8,8,8,14,14]
    for i, w in enumerate(widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

    wb.save(filepath)


# ── 导出表格数据 Excel ──────────────────────────────────────────────
# 前端 row dict key → COLUMN_NAMES 列名的映射
_COL_KEY_MAP = {
    "PartName": "part_name", "PartNo": "part_no",
    "Glass1": "glass1", "Glass2": "glass2", "Glass3": "glass3",
    "T1": "T1", "T2": "T2", "T3": "T3",
    "R1": "R1", "R2": "R2", "R3": "R3", "R4": "R4",
    "MD1": "MD1", "MD2": "MD2", "MD3": "MD3",
    "AD1": "AD1", "AD2": "AD2", "AD3": "AD3", "AD4": "AD4",
    "SavePdfFolder": "save_pdf_folder", "MfrPdfFolder": "mfr_pdf_folder",
}

_NUMERIC_EXPORT_COLS = {
    "T1", "T2", "T3", "R1", "R2", "R3", "R4",
    "MD1", "MD2", "MD3", "AD1", "AD2", "AD3", "AD4",
}


def export_batch_excel(filepath: str, rows_data):
    """
    将当前表格数据导出为可重新导入的 Excel 文件。
    列顺序与 COLUMN_NAMES 一致，确保可通过 read_excel() 重新导入。
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lens Data"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side("thin"), right=Side("thin"),
                  top=Side("thin"), bottom=Side("thin"))

    # 表头
    for i, name in enumerate(COLUMN_NAMES):
        c = ws.cell(row=1, column=i + 1, value=name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin

    # 数据行
    for r_idx, row in enumerate(rows_data, start=2):
        for c_idx, col_name in enumerate(COLUMN_NAMES, start=1):
            key = _COL_KEY_MAP.get(col_name, col_name)
            val = row.get(key, "")
            # 数值列转换
            if col_name in _NUMERIC_EXPORT_COLS:
                val = _safe_float(val)
            elif isinstance(val, str) and val.strip() == "":
                val = None
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.border = thin
            c.alignment = Alignment(horizontal="center")

    # 列宽
    widths = [14, 14, 10, 10, 10, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 14, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w

    wb.save(filepath)
